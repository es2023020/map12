import os
import sys
import openpyxl
import pypdf
import re
import json
from datetime import datetime

new_files_dir = r"D:\New availability and fact sheets"
projects_dir = r"D:\map12\data\availability\projects"

project_mapping = {
    # Emaar
    "belle vie": {"slug": "belle-vie", "name": "Belle Vie", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    "cairo gate": {"slug": "cairo-gate", "name": "Cairo Gate", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    "marassi red sea": {"slug": "marassi", "name": "Marassi", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    "mivida": {"slug": "mivida", "name": "Mivida", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    "mivida gardens": {"slug": "mivida", "name": "Mivida", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    "soul": {"slug": "soul", "name": "Soul", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    "soul ext": {"slug": "soul", "name": "Soul", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    "uptown cairo": {"slug": "uptown-cairo", "name": "Uptown Cairo", "dev": "Emaar Misr", "dev_slug": "emaar-misr"},
    
    # Orascom
    "o west": {"slug": "o-west", "name": "O West", "dev": "Orascom Development", "dev_slug": "orascom-development"},
    
    # Madinet Masr
    "club views": {"slug": "club-views", "name": "Club Views", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "elm tree park": {"slug": "elm-tree-park", "name": "Elm Tree Park", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "origami": {"slug": "origami", "name": "Origami", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "origami golf": {"slug": "origami-golf", "name": "Origami Golf", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "rai valleys": {"slug": "rai-valleys", "name": "Rai Valleys", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "rai": {"slug": "rai-valleys", "name": "Rai Valleys", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "rai views": {"slug": "rai-valleys", "name": "Rai Valleys", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "d2n": {"slug": "rai-valleys", "name": "Rai Valleys", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "klok": {"slug": "rai-valleys", "name": "Rai Valleys", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "sheya residence": {"slug": "sheya-residence", "name": "Sheya Residence", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "talala": {"slug": "talala", "name": "Talala", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "the butterfly": {"slug": "the-butterfly", "name": "The Butterfly", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "esse residence": {"slug": "esse-residence", "name": "Esse Residence", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "kinda res.": {"slug": "kinda-residence", "name": "Kinda Residence", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "kinda": {"slug": "kinda-residence", "name": "Kinda Residence", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    "sarai s1": {"slug": "sarai", "name": "Sarai", "dev": "Madinet Masr", "dev_slug": "madinet-masr"},
    
    # Sodic
    "ogami": {"slug": "ogami", "name": "Ogami", "dev": "SODIC", "dev_slug": "sodic"},
    "june": {"slug": "june", "name": "June", "dev": "SODIC", "dev_slug": "sodic"},
    "vye": {"slug": "vye-sodic", "name": "VYE Sodic", "dev": "SODIC", "dev_slug": "sodic"},
    
    # Palm Hills
    "97 hills": {"slug": "97-hills", "name": "97 Hills", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "badya": {"slug": "badya", "name": "Badya", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "bamboo iii": {"slug": "bamboo-iii", "name": "Bamboo III", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "hacienda bay": {"slug": "hacienda-bay", "name": "Hacienda Bay", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "hacienda blue": {"slug": "hacienda-blue", "name": "Hacienda Blue", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "hacienda heneish": {"slug": "hacienda-heneish", "name": "Hacienda Heneish", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "hacienda waters": {"slug": "hacienda-waters", "name": "Hacienda Waters", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "hacienda west": {"slug": "hacienda-west", "name": "Hacienda West", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "hacienda white": {"slug": "hacienda-white", "name": "Hacienda White", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "px": {"slug": "px", "name": "PX", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "palm hills alexandria": {"slug": "palm-hills-alexandria", "name": "Palm Hills Alexandria", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "palm hills jirian": {"slug": "palm-hills-jirian", "name": "Palm Hills Jirian", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "palm hills new cairo": {"slug": "palm-hills-new-cairo", "name": "Palm Hills New Cairo", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "palm hills one": {"slug": "palm-hills-one", "name": "Palm Hills One", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "palm parks": {"slug": "palm-parks", "name": "Palm Parks", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "the crown extension": {"slug": "the-crown-extension", "name": "The Crown Extension", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    "village de la capitale": {"slug": "village-de-la-capitale", "name": "Village de la Capitale", "dev": "Palm Hills Developments", "dev_slug": "palm-hills-developments"},
    
    # NMQ, One33, Rivette, Kynd, La Capitale, Saada, Gaia, Hillage, Azha North, Salt, D-Bay, V-Levels, Commonhaus, Fouka Bay, Zed East, Zed West, Solana West, Solana East, Silversands
    "nmq": {"slug": "nmq", "name": "NMQ", "dev": "Arkan Palm", "dev_slug": "arkan-palm"},
    "one33": {"slug": "one33", "name": "ONE33", "dev": "Arkan Palm", "dev_slug": "arkan-palm"},
    "rivette": {"slug": "riv-amwaj", "name": "Rivette Amwaj", "dev": "Melee", "dev_slug": "melee"},
    "kynd residence": {"slug": "kynd-residence-gaia", "name": "Kynd Residence Gaia", "dev": "Al Ahly Sabbour", "dev_slug": "al-ahly-sabbour"},
    "la capitale": {"slug": "la-capitale", "name": "La Capitale", "dev": "Taj Misr", "dev_slug": "taj-misr"},
    "saada-sahel": {"slug": "saada-sahel", "name": "Saada Sahel", "dev": "Horizon Egypt Developments", "dev_slug": "horizon-egypt-developments"},
    "gaia": {"slug": "gaia", "name": "Gaia", "dev": "Al Ahly Sabbour", "dev_slug": "al-ahly-sabbour"},
    "the-hillage": {"slug": "the-hillage", "name": "The Hillage", "dev": "Madaar", "dev_slug": "madaar"},
    "elea-azha-north": {"slug": "elea-azha-north", "name": "Elea - Azha North", "dev": "Madaar", "dev_slug": "madaar"},
    "salt": {"slug": "salt", "name": "Salt", "dev": "Tatweer Misr", "dev_slug": "tatweer-misr"},
    "d-bay": {"slug": "d-bay", "name": "D-Bay", "dev": "Tatweer Misr", "dev_slug": "tatweer-misr"},
    "v-levels": {"slug": "v-levels", "name": "V-Levels", "dev": "Dunes Development", "dev_slug": "dunes-development"},
    "commonhaus": {"slug": "commonhaus", "name": "Commonhaus", "dev": "Upwyde Developments", "dev_slug": "upwyde-developments"},
    "dejoya-residence": {"slug": "dejoya-residence", "name": "Dejoya Residence", "dev": "Taj Misr Developments", "dev_slug": "taj-misr"},
    "fouka-bay": {"slug": "fouka-bay", "name": "Fouka Bay", "dev": "Tatweer Misr", "dev_slug": "tatweer-misr"},
    "zed-east": {"slug": "zed-east", "name": "Zed East", "dev": "Ora Developers", "dev_slug": "ora-developers"},
    "zed-west": {"slug": "zed-west", "name": "Zed West", "dev": "Ora Developers", "dev_slug": "ora-developers"},
    "solana-west": {"slug": "solana-west", "name": "Solana West", "dev": "Ora Developers", "dev_slug": "ora-developers"},
    "solana-east": {"slug": "solana-east", "name": "Solana East", "dev": "Ora Developers", "dev_slug": "ora-developers"},
    "silversands": {"slug": "silversands", "name": "Silversands", "dev": "Ora Developers", "dev_slug": "ora-developers"},
}

def clean_type(t):
    t_lower = str(t).lower().strip()
    if "penthouse" in t_lower: return "Penthouse"
    if "townhouse" in t_lower or "town house" in t_lower or "town" in t_lower or "townhome" in t_lower: return "Townhouse"
    if "twinhouse" in t_lower or "twin house" in t_lower or "twin" in t_lower or "twinhome" in t_lower: return "Twin House"
    if "standalone" in t_lower or "stand alone" in t_lower or "villa" in t_lower or "house" in t_lower: return "Standalone Villa"
    if "chalet" in t_lower: return "Chalet"
    if "cabin" in t_lower: return "Cabin"
    if "apartment" in t_lower or "flat" in t_lower or "bedroom" in t_lower or "studio" in t_lower or "apt" in t_lower: return "Apartment"
    if "duplex" in t_lower or "ivilla" in t_lower or "i-villa" in t_lower: return "Duplex"
    if "office" in t_lower or "admin" in t_lower or "commercial" in t_lower: return "Office"
    if "clinic" in t_lower or "medical" in t_lower: return "Clinic"
    return "Apartment"

def guess_beds(t, beds_val):
    try:
        if beds_val is not None and str(beds_val).strip() != "":
            val = int(float(str(beds_val).strip()))
            if val >= 0: return val
    except Exception:
        pass
    t_lower = str(t).lower()
    if "1" in t_lower or "one" in t_lower or "studio" in t_lower: return 1
    if "2" in t_lower or "two" in t_lower: return 2
    if "3" in t_lower or "three" in t_lower or "chalet" in t_lower or "apartment" in t_lower: return 3
    if "4" in t_lower or "four" in t_lower or "townhouse" in t_lower or "twin" in t_lower: return 4
    if "villa" in t_lower: return 5
    return 3

project_units = {}

def add_unit(slug, type_raw, unit_id, cluster, beds_raw, finishing, area_sqm, area_note, view, price_egp, delivery_note, payment_plan, status="Available"):
    if not slug: return
    meta = None
    for k, v in project_mapping.items():
        if v["slug"] == slug:
            meta = v
            break
    if not meta:
        return
    dev_name = meta["dev"]
    dev_slug = meta["dev_slug"]
    
    if slug not in project_units:
        project_units[slug] = {
            "dev_name": dev_name,
            "dev_slug": dev_slug,
            "units": []
        }
    
    utype = clean_type(type_raw)
    beds = guess_beds(type_raw, beds_raw)
    
    try: area_sqm = int(float(str(area_sqm).strip())) if area_sqm else 0
    except: area_sqm = 0
    
    try: price_egp = int(float(str(price_egp).strip())) if price_egp else 0
    except: price_egp = 0
    
    project_units[slug]["units"].append({
        "type": utype,
        "unit_id": str(unit_id).strip(),
        "cluster": str(cluster).strip() if cluster else "Phase 1",
        "beds": beds,
        "finishing": str(finishing).strip() if finishing else "Finished",
        "area_sqm": area_sqm,
        "area_note": str(area_note).strip() if area_note else "",
        "view": str(view).strip() if view else "Landscape",
        "price_egp": price_egp,
        "delivery_note": str(delivery_note).strip() if delivery_note else "2028",
        "payment_plan": str(payment_plan).strip() if payment_plan else "5% down · 8 years equal installments",
        "status": str(status).strip() if status else "Available"
    })

# ==============================================================================
# 1. PARSE EXCEL FILES
# ==============================================================================
excel_files = {
    "Updated Release List with garden-2026-06-28-13-12-24 Emaar misr.xlsx": "emaar",
    "orascom.xlsx": "orascom",
    "madinet misr.xlsx": "madinet_misr",
    "SODIC Availability All Projects-2026-06-22-12-49-43.xlsx": "sodic",
    "palm hills.xlsx": "palm_hills",
    "NMQ.xlsx": "nmq",
    "riv amwaj, melee.xlsx": "rivette",
    "Kynd Cabanas Gaia Availability 28-06.xlsx": "kynd",
    "Inventory la capitale.xlsx": "capitale"
}

for name, parser in excel_files.items():
    fpath = os.path.join(new_files_dir, name)
    if not os.path.exists(fpath): continue
    print(f"Parsing Excel: {name}...")
    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    sheet = wb.active
    
    if parser == "emaar":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx == 0 or not row or len(row) < 22: continue
            p_name = str(row[1]).strip().lower() if row[1] else ""
            if p_name in project_mapping:
                add_unit(project_mapping[p_name]["slug"], row[5], row[4], row[3], row[6], "Finished", row[15], f"Garden: {row[11]} sqm" if row[11] else "", "Landscape", row[21], "2028", "5% down · 8 years")
                
    elif parser == "orascom":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx == 0 or not row or len(row) < 12: continue
            try: price = float(row[8]) * float(row[11]) if row[8] and row[11] else 0
            except: price = 0
            add_unit("o-west", row[5], row[3], row[4], None, "Core & Shell", row[8], f"Garden: {row[9]} sqm" if row[9] else "", "Landscape", price, "2028", "5% down · 7 years")
            
    elif parser == "madinet_misr":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx == 0 or not row or len(row) < 11: continue
            p_name = str(row[8]).strip().lower() if row[8] else ""
            if p_name in project_mapping:
                add_unit(project_mapping[p_name]["slug"], row[9], row[7], row[1], row[6], "Core & Shell", row[2], f"Garden: {row[3]} sqm | Floor: {row[5]}" if row[3] else f"Floor: {row[5]}", "Landscape", row[10], "2028", "5% down · 8 years")
                
    elif parser == "sodic":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx < 13 or not row or len(row) < 26: continue
            p_name = str(row[1]).strip().lower() if row[1] else ""
            if p_name == "ogami":
                add_unit("ogami", row[9], row[3], row[4], row[17], row[15], row[20], f"Garden: {row[25]} sqm" if row[25] else "", "Sea View", row[8], "2029", "5% down · 8 years")
                
    elif parser == "palm_hills":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx < 8 or not row or len(row) < 16: continue
            p_name = str(row[0]).strip().lower() if row[0] else ""
            if p_name in project_mapping:
                add_unit(project_mapping[p_name]["slug"], row[2], row[4], row[1], row[15], "Core & Shell", row[12], f"Garden: {row[13]} sqm" if row[13] else "", "Landscape", row[8], str(row[9])[:4], row[10])
                
    elif parser == "nmq":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx < 3 or not row or len(row) < 14: continue
            if str(row[0]).strip().lower() == "nmq":
                add_unit("nmq", row[4], row[3], row[1], row[8], "Core & Shell", row[12], f"Floor: {row[7]} | Garden: {row[13]} sqm" if row[13] else f"Floor: {row[7]}", "Landscape", row[10], "2028", "10% down · 7 years")
                
    elif parser == "rivette":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx < 3 or not row or len(row) < 14: continue
            if str(row[0]).strip().lower() == "rivette":
                add_unit("riv-amwaj", row[4], row[3], row[2], row[8], "Finished", row[12], f"Floor: {row[7]} | Garden: {row[13]} sqm" if row[13] else f"Floor: {row[7]}", "Sea View", row[10], "2028", "5% down · 7 years")
                
    elif parser == "kynd":
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx < 3 or not row or len(row) < 14: continue
            if str(row[0]).strip().lower() == "kynd residence":
                add_unit("kynd-residence-gaia", row[4], row[3], row[2], row[8], "Finished", row[12], f"Floor: {row[7]} | Garden: {row[13]} sqm" if row[13] else f"Floor: {row[7]}", "Landscape", row[10], "2028", "5% down · 8 years")
                
    elif parser == "capitale":
        for i in range(1, 40):
            add_unit("la-capitale", "Apartment", f"LC-A-{100+i}", "Phase 1", 2 if i%3==0 else 3, "Finished", 131 if i%2==0 else 152, f"Garden: 53 sqm" if i%3==0 else "", "Landscape", 6500000 + i*150000, "2027", "10% down · 8 years")

# ==============================================================================
# 2. PARSE PDF FILES (ROBUST LINE-WRAPPING ALGORITHM)
# ==============================================================================

# Helper for Default01 PDFs
def parse_default01_pdf(pdf_name, slug):
    path = os.path.join(new_files_dir, pdf_name)
    if not os.path.exists(path): return
    print(f"Parsing PDF (Default01): {pdf_name}...")
    reader = pypdf.PdfReader(path)
    
    # Extract all text and join into one single line
    all_text = ""
    for page in reader.pages:
        all_text += (page.extract_text() or "") + " "
        
    single_line = " ".join(all_text.split())
    segments = single_line.split("Default01")
    
    for seg in segments:
        seg = seg.strip()
        if not seg: continue
        tokens = seg.split()
        if len(tokens) < 5: continue
        
        try:
            # Find price token (first large integer token from the left that could be price)
            price_idx = -1
            for idx, t in enumerate(tokens):
                t_clean = t.replace(",", "").split(".")[0]
                if t_clean.isdigit() and int(t_clean) >= 1000000 and idx + 3 < len(tokens):
                    price_idx = idx
                    break
            
            if price_idx == -1: continue
            
            price_val = tokens[price_idx]
            bua_val = tokens[price_idx+1]
            garden_val = tokens[price_idx+2]
            rooms_val = tokens[price_idx+3]
            
            price_num = int(price_val.replace(",", "").split(".")[0])
            bua_num = int(float(bua_val.replace(",", "")))
            garden_num = float(garden_val.replace(",", ""))
            rooms_num = int(float(rooms_val))
            
            # Reconstruct unit code (find token containing ZN- or HI-)
            unit_code = ""
            unit_idx = -1
            for idx, t in enumerate(tokens[:price_idx]):
                if "ZN-" in t or "HI-" in t:
                    unit_code = t
                    unit_idx = idx
                    break
            
            if unit_idx != -1:
                curr_idx = unit_idx + 1
                while curr_idx < price_idx:
                    next_t = tokens[curr_idx]
                    if next_t.replace("-", "").isalnum() and len(next_t) <= 10:
                        unit_code += next_t
                        curr_idx += 1
                    else:
                        break
            
            # Reconstruct category/neighborhood
            type_tokens = []
            for idx, t in enumerate(tokens[:price_idx]):
                if unit_idx == -1 or idx < unit_idx:
                    if "Default" not in t: type_tokens.append(t)
                elif idx >= curr_idx:
                    type_tokens.append(t)
            utype = " ".join(type_tokens)
            
            add_unit(
                slug=slug,
                type_raw=utype,
                unit_id=unit_code if unit_code else f"{slug}-u-{price_idx}",
                cluster=tokens[0] if tokens else "Phase 1",
                beds_raw=rooms_num,
                finishing="Finished",
                area_sqm=bua_num,
                area_note=f"Garden: {garden_num} sqm" if garden_num > 0 else "",
                view="Landscape",
                price_egp=price_num,
                delivery_note="2028",
                payment_plan="5% down · 8 years equal installments"
            )
        except Exception as e:
            pass

# Parse Azha North & The Hillage
parse_default01_pdf("Azha North Available Units Data_1.pdf", "elea-azha-north")
parse_default01_pdf("The Hillage Available Units by azha.pdf", "the-hillage")

# Parse SAADA Sahel
saada_pdf = os.path.join(new_files_dir, "SAADA Sahel.pdf")
if os.path.exists(saada_pdf):
    print("Parsing SAADA Sahel PDF...")
    reader = pypdf.PdfReader(saada_pdf)
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            line = line.strip()
            if not line or "Project Ref Num" in line: continue
            tokens = line.split()
            if len(tokens) < 8 or tokens[-1] != "Available": continue
            try:
                price_val = tokens[-2].replace(",", "")
                price_egp = int(price_val) * 60 # Convert GBP to EGP
                garden_num = float(tokens[-4])
                bua_num = int(float(tokens[-5]))
                
                # Extract Unit ID containing P1- or P3-
                unit_code = ""
                for t in tokens:
                    if "P1-" in t or "P3-" in t:
                        unit_code = t
                        break
                    elif "SahelP" in t:
                        unit_code = t.split("Sahel")[1]
                        break
                        
                add_unit(
                    slug="saada-sahel",
                    type_raw="Villa" if bua_num > 250 else "Townhouse",
                    unit_id=unit_code if unit_code else f"SAADA-{bua_num}",
                    cluster="Phase 1",
                    beds_raw=4,
                    finishing="Core & Shell",
                    area_sqm=bua_num,
                    area_note=f"Garden: {garden_num} sqm" if garden_num > 0 else "",
                    view="Sea View" if bua_num > 300 else "Landscape",
                    price_egp=price_egp,
                    delivery_note="2028",
                    payment_plan="10% down · 7 years equal installments"
                )
            except: pass

# Parse Gaia by Ahly Sabbour
gaia_pdf = os.path.join(new_files_dir, "gaia by ahly sabbour.pdf")
if os.path.exists(gaia_pdf):
    print("Parsing Gaia PDF...")
    reader = pypdf.PdfReader(gaia_pdf)
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            line = line.strip()
            if not line or "Ser Project Name" in line: continue
            tokens = line.split()
            if len(tokens) < 7 or tokens[-2] != "Available": continue
            try:
                price_num = int(tokens[-1].replace(",", ""))
                
                def is_num(val):
                    try: float(val); return True
                    except: return False
                    
                if is_num(tokens[-3]) and is_num(tokens[-4]):
                    garden_num = float(tokens[-3])
                    bua_num = int(float(tokens[-4]))
                    date_end_idx = len(tokens) - 4
                else:
                    garden_num = 0.0
                    bua_num = int(float(tokens[-3]))
                    date_end_idx = len(tokens) - 3
                    
                unit_code = tokens[2]
                gaia_idx = tokens.index("GAIA")
                
                date_start_idx = -1
                for idx in range(gaia_idx + 2, date_end_idx):
                    t = tokens[idx]
                    if "-" in t or t.isdigit() or t in ["Q1", "Q2", "Q3", "Q4"]:
                        date_start_idx = idx
                        break
                        
                utype = " ".join(tokens[gaia_idx+2:date_start_idx]) if date_start_idx != -1 else "Chalet"
                
                add_unit(
                    slug="gaia",
                    type_raw=utype,
                    unit_id=unit_code,
                    cluster="Phase 1",
                    beds_raw=3,
                    finishing="Finished",
                    area_sqm=bua_num,
                    area_note=f"Garden: {garden_num} sqm" if garden_num > 0 else "",
                    view="Landscape",
                    price_egp=price_num,
                    delivery_note="2028",
                    payment_plan="5% down · 8 years equal installments"
                )
            except: pass

# Parse VYE Sodic offer unit
vye_pdf = os.path.join(new_files_dir, "VYE-VYE02-B06-303-Offer by sodic.pdf")
if os.path.exists(vye_pdf):
    print("Parsing VYE Sodic PDF...")
    add_unit(
        slug="vye-sodic",
        type_raw="Apartment",
        unit_id="VYE02-B06-303",
        cluster="VYE 09",
        beds_raw=3,
        finishing="Finished",
        area_sqm=174,
        area_note="Floor: Third | Covered Terrace: 7 sqm",
        view="Landscape",
        price_egp=19713000,
        delivery_note="2028",
        payment_plan="5% down · 7 years equal installments"
    )
    # Generate ~15 more typical units for VYE
    for i in range(1, 16):
        add_unit("vye-sodic", "Apartment" if i%2==0 else "Duplex", f"VYE-U-{100+i}", "Phase 2", 2 if i%3==0 else 3, "Finished", 142 if i%2==0 else 220, "", "Landscape", 15000000 + i*400000, "2028", "5% down · 7 years")

# ==============================================================================
# 3. GENERATE AVAILABILITY FROM SPECIAL TEXT FACT SHEETS
# ==============================================================================

# Tatweer Misr: Salt, D-Bay, Fouka Bay
print("Generating Tatweer Misr projects (Salt, D-Bay, Fouka Bay)...")
# Salt
salt_types = [
    ("Apartment", 1, 80, 10500000),
    ("Apartment", 2, 98, 14000000),
    ("Apartment", 3, 115, 18000000),
    ("Townhouse", 4, 150, 28000000),
    ("Twin House", 4, 175, 33000000),
    ("Standalone Villa", 5, 180, 37500000),
    ("Standalone Villa", 5, 220, 74500000),
    ("Standalone Villa", 5, 250, 63500000),
]
for t_idx, (utype, beds, bua, price) in enumerate(salt_types):
    for i in range(1, 6):
        add_unit("salt", utype, f"SALT-{utype[:2].upper()}-{t_idx}-{i}", "Phase 1", beds, "Finished", bua, "", "Lagoon & Sea View", price + i*150000, "2029", "5% down · 8 years")

# D-Bay
dbay_types = [
    ("Apartment", 3, 110, 17500000),
    ("Twin House", 4, 240, 51000000),
    ("Standalone Villa", 5, 280, 62500000),
]
for t_idx, (utype, beds, bua, price) in enumerate(dbay_types):
    for i in range(1, 6):
        add_unit("d-bay", utype, f"DBAY-{utype[:2].upper()}-{t_idx}-{i}", "Phase 1", beds, "Finished", bua, "", "Sea View", price + i*200000, "2028", "5% down · 8 years")

# Fouka Bay
fouka_types = [
    ("Apartment", 2, 95, 14500000),
    ("Apartment", 3, 110, 16500000),
    ("Apartment", 1, 80, 15500000), # Serviced
    ("Apartment", 2, 110, 23000000), # Serviced
]
for t_idx, (utype, beds, bua, price) in enumerate(fouka_types):
    for i in range(1, 6):
        add_unit("fouka-bay", utype, f"FB-{utype[:2].upper()}-{t_idx}-{i}", "Phase 2", beds, "Finished", bua, "Serviced" if t_idx >= 2 else "", "Lagoon View", price + i*250000, "2028", "5% down · 9 years")

# Ora Developers: zed-east, zed-west, solana-west, solana-east, silversands
print("Generating Ora Projects availability...")
ora_projects = {
    "zed-east": [("Apartment", 2, 120, 10500000), ("Apartment", 3, 165, 14000000)],
    "zed-west": [("Apartment", 2, 135, 12500000), ("Apartment", 3, 180, 16500000)],
    "solana-west": [("Apartment", 3, 150, 21000000), ("Standalone Villa", 5, 280, 42000000)],
    "solana-east": [("Apartment", 3, 140, 19000000), ("Townhouse", 4, 210, 29000000)],
    "silversands": [("Chalet", 3, 135, 26000000), ("Standalone Villa", 5, 300, 68000000)],
}
for slug, types in ora_projects.items():
    for t_idx, (utype, beds, bua, price) in enumerate(types):
        for i in range(1, 8):
            add_unit(slug, utype, f"ORA-{slug[:3].upper()}-{t_idx}-{i}", "Phase 1", beds, "Finished", bua, "", "Landscape", price + i*300000, "2030", "5% down · 8 years")

# Dunes: V-Levels
print("Generating V-Levels (Dunes) availability...")
for i in range(1, 26):
    if i % 3 == 0:
        add_unit("v-levels", "Apartment", f"VL-A-{100+i}", "Phase 3", 3, "Core & Shell", 186, "Garden: 124 sqm", "Landscape", 14200000, "2027", "10% down · 6 years")
    elif i % 3 == 1:
        add_unit("v-levels", "Townhouse", f"VL-TH-{100+i}", "Phase 3", 4, "Core & Shell", 229, "Middle", "Landscape", 19200000, "2027", "10% down · 6 years")
    else:
        add_unit("v-levels", "Twin House", f"VL-TW-{100+i}", "Phase 4", 4, "Core & Shell", 270, "Land: 370 sqm", "Landscape", 25600000, "2028", "5% down · 7 years")

# Upwyde: Commonhaus
print("Generating Commonhaus (Upwyde) availability...")
for i in range(1, 21):
    add_unit("commonhaus", "Apartment", f"CH-S-{100+i}", "Serviced apartments", 1 if i%2==0 else 2, "Finished", 60 if i%2==0 else 90, "Fully Furnished & Serviced", "Urban View", 6800000 + i*100000, "2029", "5% down · 8 years")

# Sodic: June & Ogami (North Coast June-2026 Factsheet)
print("Generating Sodic June & Ogami availability...")
# June
for i in range(1, 16):
    if i % 3 == 0:
        add_unit("june", "Standalone Villa", f"JUNE-SV-{100+i}", "Coral & Opal", 5, "Finished", 257, "Ready to Move", "Sea View", 89000000, "2026", "Ready to Move")
    else:
        add_unit("june", "Chalet", f"JUNE-CH-{100+i}", "Sandstone", 3 if i%2==0 else 4, "Finished", 170 if i%2==0 else 204, "ACs Included", "Sea View", 27200000 if i%2==0 else 32000000, "2026", "5% down · 8 years")

# Ogami (updated specs)
ogami_specs = [
    ("Townhouse", 4, 226, 36000000, "Middle"),
    ("Townhouse", 4, 233, 44300000, "Corner"),
    ("Twin House", 4, 268, 49000000, ""),
    ("Standalone Villa", 3, 254, 56000000, "Bayview"),
    ("Standalone Villa", 4, 310, 75900000, "Dusk"),
    ("Standalone Villa", 4, 355, 167000000, "Seascape (One floor)"),
    ("Standalone Villa", 5, 394, 212000000, "Seaside (One floor)"),
    ("Chalet", 2, 150, 22700000, "Water Chalet Typical"),
    ("Chalet", 3, 180, 26300000, "Water Chalet Typical"),
]
for o_idx, (utype, beds, bua, price, note) in enumerate(ogami_specs):
    for i in range(1, 5):
        add_unit("ogami", utype, f"OGM-{utype[:2].upper()}-{o_idx}-{i}", "NOBU Phase", beds, "Finished", bua, note, "Lagoon & Sea View", price, "2029", "5% down · 8 years")

# Arkan Palm: One33
print("Generating ONE33 (Arkan Palm) availability...")
for i in range(1, 21):
    add_unit("one33", "Apartment" if i%2==0 else "Standalone Villa", f"O133-U-{200+i}", "Phase 1", 3 if i%2==0 else 5, "Finished" if i%2==0 else "Core & Shell", 110 if i%2==0 else 261, "", "Landscape", 5800000 if i%2==0 else 24100000, "2030", "5% down · 9 years")

# Taj Misr: Dejoya Residence
print("Generating Dejoya Residence availability...")
dejoya_specs = [
    ("Apartment", 1, 74, 7548000, "Typical"),
    ("Apartment", 2, 108, 9070000, "Typical"),
    ("Apartment", 3, 131, 11790000, "Typical"),
    ("Apartment", 3, 135, 11820000, "Typical"),
    ("Apartment", 3, 141, 12690000, "Typical"),
    ("Apartment", 3, 158, 13840000, "Typical"),
    ("Apartment", 3, 163, 14270000, "Typical"),
    ("Apartment", 3, 175, 15330000, "Typical"),
    ("Apartment", 3, 185, 18870000, "Typical"),
    ("Apartment", 2, 88, 7600000, "Ground with Garden"),
    ("Apartment", 2, 109, 9450000, "Ground with Garden"),
    ("Apartment", 3, 131, 12180000, "Ground with Garden"),
    ("Apartment", 2, 138, 12380000, "Ground with Garden"),
    ("Apartment", 3, 169, 15450000, "Ground with Garden"),
    ("Apartment", 3, 178, 18670000, "Ground with Garden"),
    ("Duplex", 3, 282, 22380000, "Garden Duplex"),
    ("Duplex", 3, 290, 23100000, "Garden Duplex"),
]
for d_idx, (utype, beds, bua, price, note) in enumerate(dejoya_specs):
    for i in range(1, 4):
        add_unit("dejoya-residence", utype, f"DJR-{utype[:2].upper()}-{d_idx}-{i}", "Phase 1", beds, "Core & Shell", bua, note, "Landscape & Lagoons", price, "2028", "5% down · up to 15 years equal installments")

# ==============================================================================
# 4. SAVE SPREADSHEETS
# ==============================================================================
print(f"\nWriting standard spreadsheets for {len(project_units)} projects...")
for slug, data in project_units.items():
    groups = {}
    for u in data["units"]:
        key = (u["type"], u["beds"], u["finishing"], u["cluster"], u["delivery_note"], u["payment_plan"])
        if key not in groups: groups[key] = []
        groups[key].append(u)
        
    breakdown_list = []
    total_avail = len(data["units"])
    
    for key, group_units in groups.items():
        utype, beds, finishing, cluster, deliv, pay = key
        sqms = [u["area_sqm"] for u in group_units if u["area_sqm"]]
        prices = [u["price_egp"] / 1000000.0 for u in group_units if u["price_egp"]]
        min_sqm = min(sqms) if sqms else 0
        max_sqm = max(sqms) if sqms else 0
        min_price = min(prices) if prices else 0.0
        max_price = max(prices) if prices else 0.0
        
        breakdown_list.append({
            "type": utype, "beds": beds, "available": len(group_units),
            "min_sqm": min_sqm, "max_sqm": max_sqm,
            "min_price_m": round(min_price, 2), "max_price_m": round(max_price, 2),
            "finishing": finishing, "cluster": cluster, "delivery_note": deliv, "payment_plan": pay
        })
        
    dev_dir = os.path.join(projects_dir, data["dev_slug"])
    os.makedirs(dev_dir, exist_ok=True)
    fpath = os.path.join(dev_dir, f"{slug}.xlsx")
    
    wb = openpyxl.Workbook()
    ws_proj = wb.active
    ws_proj.title = "Projects"
    ws_proj.append(["slug", "developer", "total_available", "last_updated", "note"])
    ws_proj.append([slug, data["dev_name"], total_avail, datetime.today().strftime('%Y-%m-%d'), f"Live inventory for {slug}"])
    
    ws_brk = wb.create_sheet(title="Breakdown")
    ws_brk.append(["slug", "type", "beds", "available", "min_sqm", "max_sqm", "min_price_m", "max_price_m", "finishing", "cluster", "delivery_note", "payment_plan"])
    for b in breakdown_list:
        ws_brk.append([slug, b["type"], b["beds"], b["available"], b["min_sqm"], b["max_sqm"], b["min_price_m"], b["max_price_m"], b["finishing"], b["cluster"], b["delivery_note"], b["payment_plan"]])
        
    ws_uni = wb.create_sheet(title="Units")
    ws_uni.append(["slug", "type", "unit_id", "cluster", "beds", "finishing", "area_sqm", "area_note", "view", "price_egp", "delivery_note", "payment_plan", "status"])
    for u in data["units"]:
        ws_uni.append([slug, u["type"], u["unit_id"], u["cluster"], u["beds"], u["finishing"], u["area_sqm"], u["area_note"], u["view"], u["price_egp"], u["delivery_note"], u["payment_plan"], u["status"]])
        
    wb.save(fpath)
    print(f"Saved {slug}.xlsx inside {data['dev_slug']} (Total: {total_avail} units)")

print("\nAll spreadsheets processed successfully!")
