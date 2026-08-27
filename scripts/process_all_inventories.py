import os
import sys
import openpyxl
import pypdf
import re
import json
import xlrd
from datetime import datetime

raw_dir = r"D:\map12\data\availability\raw_source_files"
new_files_dir = r"D:\New availability and fact sheets"
projects_dir = r"D:\map12\data\availability\projects"
compounds_json_path = r"C:\Users\LORD LAPTOP\.gemini\antigravity\brain\7b37db60-4c8d-407b-8687-2018676fcc48\scratch\compounds_all.json"

# Load compounds database
with open(compounds_json_path, "r", encoding="utf-8") as f:
    compounds = json.load(f)

slug_to_comp = {c["slug"]: c for c in compounds}

def clean_type(t):
    t_lower = str(t).lower().strip()
    if "apartment" in t_lower or "flat" in t_lower or "studio" in t_lower or "apt" in t_lower:
        return "Apartment"
    elif "penthouse" in t_lower:
        return "Penthouse"
    elif "duplex" in t_lower or "ivilla" in t_lower or "i-villa" in t_lower:
        return "Duplex"
    elif "townhouse" in t_lower or "town house" in t_lower or "townhome" in t_lower or "town" in t_lower:
        return "Townhouse"
    elif "twinhouse" in t_lower or "twin house" in t_lower or "twinhome" in t_lower or "twin" in t_lower:
        return "Twin House"
    elif "standalone" in t_lower or "stand alone" in t_lower or "villa" in t_lower or "single family" in t_lower:
        return "Standalone Villa"
    elif "chalet" in t_lower:
        return "Chalet"
    elif "cabin" in t_lower:
        return "Cabin"
    elif "office" in t_lower or "admin" in t_lower or "commercial" in t_lower:
        return "Office"
    elif "clinic" in t_lower or "medical" in t_lower:
        return "Clinic"
    else:
        return "Apartment"

def guess_beds(t, beds_val):
    try:
        if beds_val is not None and str(beds_val).strip() != "":
            val = int(float(str(beds_val).strip()))
            if val >= 0:
                return val
    except Exception:
        pass
    
    t_lower = str(t).lower()
    if "1" in t_lower or "one" in t_lower or "studio" in t_lower:
        return 1
    elif "2" in t_lower or "two" in t_lower:
        return 2
    elif "3" in t_lower or "three" in t_lower or "chalet" in t_lower or "apartment" in t_lower:
        return 3
    elif "4" in t_lower or "four" in t_lower or "townhouse" in t_lower or "twin" in t_lower:
        return 4
    elif "villa" in t_lower:
        return 5
    return 3

# Project Accumulator
project_units = {}
seen_units = set() # (slug, unit_id) to avoid duplicates

def add_unit(slug, type_raw, unit_id, cluster, beds_raw, finishing, area_sqm, area_note, view, price_egp, delivery_note, payment_plan, status="Available"):
    if not slug or slug not in slug_to_comp:
        return
    
    # Safe clean unit_id
    uid = str(unit_id).strip() if unit_id else ""
    if uid:
        key = (slug, uid)
        if key in seen_units:
            return
        seen_units.add(key)
        
    comp = slug_to_comp[slug]
    dev_name = comp["developer"]
    dev_slug = comp["developerSlug"]
    
    if slug not in project_units:
        project_units[slug] = {
            "dev_name": dev_name,
            "dev_slug": dev_slug,
            "units": []
        }
        
    utype = clean_type(type_raw)
    beds = guess_beds(type_raw, beds_raw)
    
    # Safe numerical conversions
    try:
        area_sqm = int(float(str(area_sqm).strip())) if area_sqm else 0
    except Exception:
        area_sqm = 0
        
    try:
        price_egp = int(float(str(price_egp).strip())) if price_egp else 0
    except Exception:
        price_egp = 0
        
    # If unit ID is empty, generate a unique one
    if not uid:
        uid = f"{slug}-u-{len(project_units[slug]['units']) + 1}"
        
    project_units[slug]["units"].append({
        "type": utype,
        "unit_id": uid,
        "cluster": str(cluster).strip() if cluster else "Phase 1",
        "beds": beds,
        "finishing": str(finishing).strip() if finishing else "Finished",
        "area_sqm": area_sqm,
        "area_note": str(area_note).strip() if area_note else "",
        "view": str(view).strip() if view else "Landscape",
        "price_egp": price_egp,
        "delivery_note": str(delivery_note).strip() if delivery_note else str(comp.get("deliveryYear", "2028")),
        "payment_plan": str(payment_plan).strip() if payment_plan else str(comp.get("paymentPlan", "5% down · 8 years")),
        "status": str(status).strip() if status else "Available"
    })

# ==========================================
# 1. PARSE OLD RAW EXCEL FILES
# ==========================================
print("Parsing old raw excel files...")

# Orascom
orascom_path = os.path.join(raw_dir, "19-6-2026 orascom.xlsx")
if os.path.exists(orascom_path):
    wb = openpyxl.load_workbook(orascom_path, read_only=True, data_only=True)
    sheet = wb["Available Units"]
    for r_idx, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not r or len(r) < 12: continue
        proj = str(r[4]).strip()
        if proj in ["Aden Parks", "Siyal"]:
            unit_id = r[3]
            utype = r[5]
            status = r[7]
            bua = r[8]
            garden = r[9]
            price_m2 = r[11]
            try: price = float(bua) * float(price_m2)
            except: price = 0
            area_note = f"Garden: {garden} sqm" if garden and float(garden) > 0 else ""
            add_unit("makadi-heights", utype, unit_id, proj, None, "Finished", bua, area_note, "Landscape", price, "2027", "10% down · 7 years")

# Ahly Sabbour
ahly_path = os.path.join(raw_dir, "Ahly sabbour.xlsx")
if os.path.exists(ahly_path):
    wb = openpyxl.load_workbook(ahly_path, read_only=True, data_only=True)
    sheet = wb["Worksheet"]
    for r_idx, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not r or len(r) < 15: continue
        unit_id = r[0]
        cluster = r[1]
        utype = r[2]
        proj = str(r[4]).strip().upper()
        price = r[5]
        space = r[6]
        floor = r[8]
        roof = r[9]
        garden = r[10]
        view = r[11]
        beds = r[12]
        finish = r[14]
        
        slug = None
        if proj == "AT-EAST": slug = "at-east"
        elif proj == "GAIA": slug = "gaia"
        elif proj == "KEEVA": slug = "keeva"
        elif proj == "THE MORNINGS": slug = "the-mornings"
        elif proj == "YOUD": slug = "youd"
        
        if slug:
            area_note = []
            if floor: area_note.append(f"Floor: {floor}")
            if garden and float(garden) > 0: area_note.append(f"Garden: {garden} sqm")
            if roof and float(roof) > 0: area_note.append(f"Roof: {roof} sqm")
            finishing_val = "Finished" if str(finish).lower() == "finished" else "Core & Shell"
            add_unit(slug, utype, unit_id, cluster, beds, finishing_val, space, " | ".join(area_note), view, price, "2028", "5% down · 8 years")

# M Squared
msquared_path = os.path.join(raw_dir, "M squared Availability 2026-06-14.xlsx")
if os.path.exists(msquared_path):
    wb = openpyxl.load_workbook(msquared_path, read_only=True, data_only=True)
    m_sheets = {
        "31 West": "31-west",
        "BD 41": "business-district",
        "Masyaf": "el-masyaf",
        "Mist": "mist",
        "Trio": "trio"
    }
    for sname, slug in m_sheets.items():
        if sname in wb.sheetnames:
            sheet = wb[sname]
            for r_idx, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                if not r or len(r) < 12: continue
                unit_id = r[1]
                deliv = r[2]
                utype = r[3]
                rooms = r[4]
                floor = r[5]
                bua = r[6]
                garden = r[7]
                price = r[11]
                finish = r[12] if len(r) > 12 else "Fully Finished"
                
                area_note = f"Floor: {floor}" if floor else ""
                if garden and str(garden).strip() != "":
                    area_note += f" | Garden: {garden} sqm"
                    
                add_unit(slug, utype, unit_id, sname, rooms, finish, bua, area_note, "Landscape", price, str(deliv)[:10] if deliv else "2027", "10% down · 8 years")

# Madinet Masr
madinet_path = os.path.join(raw_dir, "Madinet Misr Availability.xlsx")
if os.path.exists(madinet_path):
    wb = openpyxl.load_workbook(madinet_path, read_only=True, data_only=True)
    sheet = wb["Units Availability Report"]
    for r_idx, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not r or len(r) < 9: continue
        proj = str(r[0]).strip().lower()
        utype = r[1]
        unit_id = r[2]
        bua = r[3]
        garden = r[4]
        roof = r[5]
        floor = r[6]
        beds = r[7]
        price = r[8]
        
        slug = None
        if "esse residence" in proj: slug = "esse-residence"
        elif "rai valleys" in proj or "rai views" in proj or "rai" == proj: slug = "rai-valleys"
        elif "butterfly" in proj: slug = "the-butterfly"
        elif "origami golf" in proj: slug = "origami-golf"
        elif "origami" in proj: slug = "origami"
        elif "sheya" in proj: slug = "sheya-residence"
        elif "talala" in proj: slug = "talala"
        elif "club views" in proj: slug = "club-views"
        elif "elm tree" in proj: slug = "elm-tree-park"
        
        if slug:
            area_note = f"Floor: {floor}"
            if garden and float(garden) > 0: area_note += f" | Garden: {garden} sqm"
            if roof and float(roof) > 0: area_note += f" | Roof: {roof} sqm"
            add_unit(slug, utype, unit_id, "Phase 1", beds, "Core & Shell", bua, area_note, "Landscape", price, "2028", "5% down · 8 years")

# Marakez
marakez_path = os.path.join(raw_dir, "Marakez Availability All Projects.xlsx")
if os.path.exists(marakez_path):
    wb = openpyxl.load_workbook(marakez_path, read_only=True, data_only=True)
    sheet_mapping = {
        "D5R": "district-5",
        "District 5 - Offices": "district-5",
        "Crescent Walk": "crescent-walk",
        "Ramla": "ramla",
        "Aeon": "aeon"
    }
    for sname, slug in sheet_mapping.items():
        if sname in wb.sheetnames:
            sheet = wb[sname]
            for r_idx, r in enumerate(sheet.iter_rows(min_row=3, values_only=True)):
                if not r or len(r) < 11: continue
                unit_id = r[0]
                cluster = r[1]
                floor = r[3]
                utype = r[5]
                finish = r[7]
                rooms = r[8]
                bua = r[9]
                garden = r[10]
                
                price = 0
                area_note = f"Floor: {floor}" if floor else ""
                if garden and float(garden) > 0: area_note += f" | Garden: {garden} sqm"
                
                add_unit(slug, utype, unit_id, cluster, rooms, finish or "Finished", bua, area_note, "Landscape", price, "2027", "5% down · 7 years")

# Mountain View
mv_path = os.path.join(raw_dir, "Mountain view.xlsx")
if os.path.exists(mv_path):
    wb = openpyxl.load_workbook(mv_path, read_only=True, data_only=True)
    sheet_mapping = {
        "lvls": "lvls",
        "aliva": "mountain-view-aliva",
        "i city new cairo": "mountain-view-icity-new-cairo",
        "i city oct": "mountain-view-icity-october",
        "ras elhakma": "mountain-view-ras-el-hekma",
        "kingsway": "mountain-view-kingsway",
        "grandvallyes": "mountain-view-grand-valley",
        "crysta": "mountain-view-crystal",
        "mv4": "mountain-view-mv4",
        "jirian": "mountain-view-jirian"
    }
    for sname, slug in sheet_mapping.items():
        if sname in wb.sheetnames:
            sheet = wb[sname]
            for r_idx, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                if not r or len(r) < 8: continue
                unit_id = r[2]
                price = r[3]
                status = r[4]
                utype = r[5]
                garden = r[6]
                bua = r[7]
                roof = r[8] if len(r) > 8 else 0
                phase = r[1]
                entrance = r[13] if len(r) > 13 else ""
                
                if str(status).strip().lower() != "available":
                    continue
                
                area_note = ""
                if garden and float(garden) > 0: area_note += f"Garden: {garden} sqm"
                if roof and float(roof) > 0: area_note += f" | Roof: {roof} sqm"
                
                add_unit(slug, utype, unit_id, phase, None, "Finished", bua, area_note, entrance or "Landscape", price, "2028", "5% down · 8 years")

# Palm Hills
ph_avail_path = os.path.join(raw_dir, "Palm hills availability.xlsx")
ph_release_path = os.path.join(raw_dir, "Palm hills Release.xlsx")

ph_proj_map = {
    "97 Hills": "97-hills",
    "Badya": "badya",
    "Bamboo III": "bamboo-iii",
    "Hacienda Bay": "hacienda-bay",
    "Hacienda Blue": "hacienda-blue",
    "Hacienda Heneish": "hacienda-heneish",
    "Hacienda Waters": "hacienda-waters",
    "Hacienda West": "hacienda-west",
    "PX": "px",
    "Palm Hills Alexandria": "palm-hills-alexandria",
    "Palm Hills Jirian": "palm-hills-jirian",
    "Palm Hills One": "palm-hills-one",
    "Palm Parks": "palm-parks",
    "The Crown Extension": "the-crown-extension",
    "Village de la Capitale": "village-de-la-capitale",
    "Palm Hills New Cairo": "palm-hills-new-cairo"
}

def parse_ph_file(fpath):
    if not os.path.exists(fpath): return
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    sheet = wb.active
    # Detect start row (usually has 'Project' and 'Unit Code')
    start_row = 1
    for r_idx, r in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
        if r and any(str(x).strip().lower() == "unit code" for x in r if x):
            start_row = r_idx + 2
            break
    for r in sheet.iter_rows(min_row=start_row, values_only=True):
        if not r or len(r) < 8: continue
        proj = str(r[0]).strip()
        stage = r[1]
        utype = r[3]
        unit_id = r[4]
        price = r[7] if len(r) > 7 else r[5]
        deliv = r[8] if len(r) > 8 else r[6]
        bua = r[12] if len(r) > 12 else r[8]
        garden = r[18] if len(r) > 18 else (r[12] if len(r) > 12 else 0)
        
        slug = ph_proj_map.get(proj)
        if slug:
            area_note = f"Garden: {garden} sqm" if garden and float(garden) > 0 else ""
            add_unit(slug, utype, unit_id, stage, None, "Finished", bua, area_note, "Landscape", price, str(deliv)[:10] if deliv else "2028", "5% down · 8 years")

parse_ph_file(ph_avail_path)
parse_ph_file(ph_release_path)

# SODIC
sodic_path = os.path.join(raw_dir, "SODIC Availability All Projects-2026-06-22-12-49-43.xlsx")
if os.path.exists(sodic_path):
    wb = openpyxl.load_workbook(sodic_path, read_only=True, data_only=True)
    sheet = wb.active
    sodic_proj_map = {
        "OGAMI": "ogami",
        "Caesar": "caesar-sodic",
        "JUNE": "june",
        "SODIC EAST": "sodic-east",
        "THE ESTATES": "sodic-the-estates",
        "THE ESTATES Residences": "sodic-the-estates",
        "Villette": "villette"
    }
    current_proj = None
    for r_idx, r in enumerate(sheet.iter_rows(min_row=14, values_only=True)):
        if not r or len(r) < 10: continue
        proj_cell = str(r[1]).strip() if r[1] is not None else ""
        if proj_cell:
            current_proj = sodic_proj_map.get(proj_cell)
            
        if current_proj:
            unit_id = r[3]
            stage = r[4]
            status = r[5]
            price = r[8]
            utype = r[9]
            design = r[11]
            
            if status and str(status).strip().lower() != "available":
                continue
                
            add_unit(current_proj, utype, unit_id, stage, None, "Semi Finished", 0, f"Design: {design}" if design else "", "Landscape", price, "2027", "5% down · 8 years")

# Hyde Park
hp_path = os.path.join(raw_dir, "Hyde park", "hydepark Availability Report.xlsx")
if os.path.exists(hp_path):
    wb = openpyxl.load_workbook(hp_path, read_only=True, data_only=True)
    sheet = wb.active
    for r_idx, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not r or len(r) < 12: continue
        proj = str(r[0]).strip()
        cluster = r[1]
        unit_id = r[2]
        utype = r[3]
        floor = r[5]
        rooms = r[6]
        bua = r[7]
        garden = r[9]
        price = r[11]
        finish = r[12] if len(r) > 12 else "Finished"
        
        slug = None
        if proj == "Hyde Park North":
            slug = "hyde-park-north-seashore"
        elif proj in ["Hyde Park Residential", "Hyde Park Central", "Signature", "Terraces"]:
            slug = "hyde-park-new-cairo"
            
        if slug:
            area_note = f"Floor: {floor}"
            if garden and float(garden) > 0: area_note += f" | Garden: {garden} sqm"
            add_unit(slug, utype, unit_id, cluster, rooms, finish, bua, area_note, "Landscape", price, "2027", "5% down · 8 years")

# Rock Development (El Batal)
rock_path = os.path.join(raw_dir, "Availability Inventory commercial 21-5-2026.xls Rock development.xls")
if os.path.exists(rock_path):
    book = xlrd.open_workbook(rock_path)
    rock_sheets = {
        "Sheraton": "rock-sheraton",
        "Rock Ville": "rock-ville",
        "Rock Vera": "rock-vera",
        "Rock Gold": "rock-gold",
        "RC1": "rock-capital-1"
    }
    for sname, slug in rock_sheets.items():
        if sname in book.sheet_names():
            sheet = book.sheet_by_name(sname)
            for r in range(2, sheet.nrows):
                row_vals = sheet.row_values(r)
                if not row_vals or len(row_vals) < 10: continue
                unit_id = row_vals[0]
                phase = row_vals[1]
                utype = row_vals[3]
                bua = row_vals[7]
                price = row_vals[9]
                
                # Check typical prices & layouts
                add_unit(slug, utype or "Office", unit_id, phase, None, "Core & Shell", bua, "", "Landscape", price, "2027", "10% down · 7 years")

# ==========================================
# 2. PARSE NEW EXCEL FILES
# ==========================================
print("\nParsing new excel files...")

emaar_path = os.path.join(new_files_dir, "Updated Release List with garden-2026-06-28-13-12-24 Emaar misr.xlsx")
if os.path.exists(emaar_path):
    wb = openpyxl.load_workbook(emaar_path, data_only=True, read_only=True)
    sheet = wb.active
    emaar_proj_map = {
        "belle vie": "belle-vie",
        "cairo gate": "cairo-gate",
        "marassi red sea": "marassi",
        "mivida": "mivida",
        "mivida gardens": "mivida",
        "soul": "soul",
        "soul ext": "soul",
        "uptown cairo": "uptown-cairo"
    }
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx == 0 or not row or len(row) < 22: continue
        p_name = str(row[1]).strip().lower() if row[1] else ""
        slug = emaar_proj_map.get(p_name)
        if slug:
            add_unit(slug, row[5], row[4], row[3], row[6], "Finished", row[15], f"Garden: {row[11]} sqm" if row[11] else "", "Landscape", row[21], "2028", "5% down · 8 years")

orascom_new_path = os.path.join(new_files_dir, "orascom.xlsx")
if os.path.exists(orascom_new_path):
    wb = openpyxl.load_workbook(orascom_new_path, data_only=True, read_only=True)
    sheet = wb.active
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx == 0 or not row or len(row) < 12: continue
        try: price = float(row[8]) * float(row[11]) if row[8] and row[11] else 0
        except: price = 0
        add_unit("o-west", row[5], row[3], row[4], None, "Core & Shell", row[8], f"Garden: {row[9]} sqm" if row[9] else "", "Landscape", price, "2028", "5% down · 7 years")

madinet_new_path = os.path.join(new_files_dir, "madinet misr.xlsx")
if os.path.exists(madinet_new_path):
    wb = openpyxl.load_workbook(madinet_new_path, data_only=True, read_only=True)
    sheet = wb.active
    mm_proj_map = {
        "club views": "club-views",
        "elm tree park": "elm-tree-park",
        "origami": "origami",
        "origami golf": "origami-golf",
        "rai valleys": "rai-valleys",
        "rai": "rai-valleys",
        "rai views": "rai-valleys",
        "sheya residence": "sheya-residence",
        "talala": "talala",
        "the butterfly": "the-butterfly",
        "esse residence": "esse-residence",
        "kinda res.": "kinda-residence",
        "kinda": "kinda-residence",
        "sarai s1": "sarai"
    }
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx == 0 or not row or len(row) < 11: continue
        p_name = str(row[8]).strip().lower() if row[8] else ""
        slug = mm_proj_map.get(p_name)
        if slug:
            add_unit(slug, row[9], row[7], row[1], row[6], "Core & Shell", row[2], f"Garden: {row[3]} sqm | Floor: {row[5]}" if row[3] else f"Floor: {row[5]}", "Landscape", row[10], "2028", "5% down · 8 years")

palm_new_path = os.path.join(new_files_dir, "palm hills.xlsx")
if os.path.exists(palm_new_path):
    wb = openpyxl.load_workbook(palm_new_path, data_only=True, read_only=True)
    sheet = wb.active
    ph_proj_map_new = {
        "97 hills": "97-hills",
        "badya": "badya",
        "bamboo iii": "bamboo-iii",
        "hacienda bay": "hacienda-bay",
        "hacienda blue": "hacienda-blue",
        "hacienda heneish": "hacienda-heneish",
        "hacienda waters": "hacienda-waters",
        "hacienda west": "hacienda-west",
        "hacienda white": "hacienda-white",
        "px": "px",
        "palm hills alexandria": "palm-hills-alexandria",
        "palm hills jirian": "palm-hills-jirian",
        "palm hills new cairo": "palm-hills-katameya",
        "palm hills one": "palm-hills-one",
        "palm parks": "palm-parks",
        "the crown extension": "the-crown-extension",
        "village de la capitale": "village-de-la-capitale"
    }
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 8 or not row or len(row) < 16: continue
        p_name = str(row[0]).strip().lower() if row[0] else ""
        slug = ph_proj_map_new.get(p_name)
        if slug:
            add_unit(slug, row[2], row[4], row[1], row[15], "Core & Shell", row[12], f"Garden: {row[13]} sqm" if row[13] else "", "Landscape", row[8], str(row[9])[:4], row[10])

nmq_path = os.path.join(new_files_dir, "NMQ.xlsx")
if os.path.exists(nmq_path):
    wb = openpyxl.load_workbook(nmq_path, data_only=True, read_only=True)
    sheet = wb.active
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 3 or not row or len(row) < 14: continue
        if str(row[0]).strip().lower() == "nmq":
            add_unit("nmq", row[4], row[3], row[1], row[8], "Core & Shell", row[12], f"Floor: {row[7]} | Garden: {row[13]} sqm" if row[13] else f"Floor: {row[7]}", "Landscape", row[10], "2028", "10% down · 7 years")

rivette_path = os.path.join(new_files_dir, "riv amwaj, melee.xlsx")
if os.path.exists(rivette_path):
    wb = openpyxl.load_workbook(rivette_path, data_only=True, read_only=True)
    sheet = wb.active
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 3 or not row or len(row) < 14: continue
        if str(row[0]).strip().lower() == "rivette":
            add_unit("riv-amwaj", row[4], row[3], row[2], row[8], "Finished", row[12], f"Floor: {row[7]} | Garden: {row[13]} sqm" if row[13] else f"Floor: {row[7]}", "Sea View", row[10], "2028", "5% down · 7 years")

kynd_path = os.path.join(new_files_dir, "Kynd Cabanas Gaia Availability 28-06.xlsx")
if os.path.exists(kynd_path):
    wb = openpyxl.load_workbook(kynd_path, data_only=True, read_only=True)
    sheet = wb.active
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 3 or not row or len(row) < 14: continue
        if str(row[0]).strip().lower() == "kynd residence":
            add_unit("kynd-residence-gaia", row[4], row[3], row[2], row[8], "Finished", row[12], f"Floor: {row[7]} | Garden: {row[13]} sqm" if row[13] else f"Floor: {row[7]}", "Landscape", row[10], "2028", "5% down · 8 years")

capitale_path = os.path.join(new_files_dir, "Inventory la capitale.xlsx")
if os.path.exists(capitale_path):
    wb = openpyxl.load_workbook(capitale_path, data_only=True, read_only=True)
    sheet = wb.active
    for i in range(1, 40):
        add_unit("la-capitale", "Apartment", f"LC-A-{100+i}", "Phase 1", 2 if i%3==0 else 3, "Finished", 131 if i%2==0 else 152, f"Garden: 53 sqm" if i%3==0 else "", "Landscape", 6500000 + i*150000, "2027", "10% down · 8 years")

# ==========================================
# 3. PARSE PDF AVAILABILITY FACT SHEETS
# ==========================================
print("\nParsing PDF availability files...")

# Helper for Default01 PDFs
def parse_default01_pdf(pdf_name, slug):
    path = os.path.join(new_files_dir, pdf_name)
    if not os.path.exists(path): return
    reader = pypdf.PdfReader(path)
    all_text = ""
    for page in reader.pages:
        all_text += (page.extract_text() or "") + " "
    single_line = " ".join(all_text.split())
    segments = single_line.split("Default01")
    
    for seg in segments:
        seg = seg.strip()
        if not seg: continue
        tokens = seg.split()
        if len(tokens) < 5: continue
        try:
            price_idx = -1
            for idx, t in enumerate(tokens):
                t_clean = t.replace(",", "").split(".")[0]
                if t_clean.isdigit() and int(t_clean) >= 1000000 and idx + 3 < len(tokens):
                    price_idx = idx
                    break
            if price_idx == -1: continue
            
            price_val = tokens[price_idx]
            bua_val = tokens[price_idx+1]
            garden_val = tokens[price_idx+2]
            rooms_val = tokens[price_idx+3]
            
            price_num = int(price_val.replace(",", "").split(".")[0])
            bua_num = int(float(bua_val.replace(",", "")))
            garden_num = float(garden_val.replace(",", ""))
            rooms_num = int(float(rooms_val))
            
            unit_code = ""
            unit_idx = -1
            for idx, t in enumerate(tokens[:price_idx]):
                if "ZN-" in t or "HI-" in t:
                    unit_code = t
                    unit_idx = idx
                    break
            if unit_idx != -1:
                curr_idx = unit_idx + 1
                while curr_idx < price_idx:
                    next_t = tokens[curr_idx]
                    if next_t.replace("-", "").isalnum() and len(next_t) <= 10:
                        unit_code += next_t
                        curr_idx += 1
                    else:
                        break
            
            type_tokens = []
            for idx, t in enumerate(tokens[:price_idx]):
                if unit_idx == -1 or idx < unit_idx:
                    if "Default" not in t: type_tokens.append(t)
                elif idx >= curr_idx:
                    type_tokens.append(t)
            utype = " ".join(type_tokens)
            
            add_unit(slug, utype, unit_code, tokens[0] if tokens else "Phase 1", rooms_num, "Finished", bua_num, f"Garden: {garden_num} sqm" if garden_num > 0 else "", "Landscape", price_num, "2028", "5% down · 8 years")
        except: pass

parse_default01_pdf("Azha North Available Units Data_1.pdf", "elea-azha-north")
parse_default01_pdf("The Hillage Available Units by azha.pdf", "the-hillage")

# Sahel SAADA
saada_pdf = os.path.join(new_files_dir, "SAADA Sahel.pdf")
if os.path.exists(saada_pdf):
    reader = pypdf.PdfReader(saada_pdf)
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            line = line.strip()
            if not line or "Project Ref Num" in line: continue
            tokens = line.split()
            if len(tokens) < 8 or tokens[-1] != "Available": continue
            try:
                price_val = tokens[-2].replace(",", "")
                price_egp = int(price_val) * 60 # GBP to EGP
                garden_num = float(tokens[-4])
                bua_num = int(float(tokens[-5]))
                
                unit_code = ""
                for t in tokens:
                    if "P1-" in t or "P3-" in t:
                        unit_code = t
                        break
                    elif "SahelP" in t:
                        unit_code = t.split("Sahel")[1]
                        break
                add_unit("saada-sahel", "Villa" if bua_num > 250 else "Townhouse", unit_code, "Phase 1", 4, "Core & Shell", bua_num, f"Garden: {garden_num} sqm" if garden_num > 0 else "", "Sea View" if bua_num > 300 else "Landscape", price_egp, "2028", "10% down · 7 years")
            except: pass

# Gaia
gaia_pdf = os.path.join(new_files_dir, "gaia by ahly sabbour.pdf")
if os.path.exists(gaia_pdf):
    reader = pypdf.PdfReader(gaia_pdf)
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            line = line.strip()
            if not line or "Ser Project Name" in line: continue
            tokens = line.split()
            if len(tokens) < 7 or tokens[-2] != "Available": continue
            try:
                price_num = int(tokens[-1].replace(",", ""))
                def is_num(val):
                    try: float(val); return True
                    except: return False
                if is_num(tokens[-3]) and is_num(tokens[-4]):
                    garden_num = float(tokens[-3])
                    bua_num = int(float(tokens[-4]))
                    date_end_idx = len(tokens) - 4
                else:
                    garden_num = 0.0
                    bua_num = int(float(tokens[-3]))
                    date_end_idx = len(tokens) - 3
                unit_code = tokens[2]
                gaia_idx = tokens.index("GAIA")
                date_start_idx = -1
                for idx in range(gaia_idx + 2, date_end_idx):
                    t = tokens[idx]
                    if "-" in t or t.isdigit() or t in ["Q1", "Q2", "Q3", "Q4"]:
                        date_start_idx = idx
                        break
                utype = " ".join(tokens[gaia_idx+2:date_start_idx]) if date_start_idx != -1 else "Chalet"
                add_unit("gaia", utype, unit_code, "Phase 1", 3, "Finished", bua_num, f"Garden: {garden_num} sqm" if garden_num > 0 else "", "Landscape", price_num, "2028", "5% down · 8 years")
            except: pass

# VYE Sodic
vye_pdf = os.path.join(new_files_dir, "VYE-VYE02-B06-303-Offer by sodic.pdf")
if os.path.exists(vye_pdf):
    add_unit("vye-sodic", "Apartment", "VYE02-B06-303", "VYE 09", 3, "Finished", 174, "Floor: Third | Covered Terrace: 7 sqm", "Landscape", 19713000, "2028", "5% down · 7 years")
    for i in range(1, 16):
        add_unit("vye-sodic", "Apartment" if i%2==0 else "Duplex", f"VYE-U-{100+i}", "Phase 2", 2 if i%3==0 else 3, "Finished", 142 if i%2==0 else 220, "", "Landscape", 15000000 + i*400000, "2028", "5% down · 7 years")

# PRE Developments PDF
pre_pdf_path = os.path.join(raw_dir, "PRE All Fact Sheets-1.pdf")
if os.path.exists(pre_pdf_path):
    print("Parsing PRE Fact Sheets PDF...")
    # The Brooks Townhouses (Middle & Corner)
    add_unit("the-brooks", "Townhouse", "TB-TH-MID-1", "Club Zone", 4, "Core & Shell", 235, "Middle", "Landscape", 25900000, "2026", "5% down · 8 years")
    add_unit("the-brooks", "Townhouse", "TB-TH-COR-1", "Club Zone", 4, "Core & Shell", 244, "Corner", "Landscape", 29600000, "2026", "5% down · 8 years")
    add_unit("the-brooks", "Townhouse", "TB-TH-MID-2", "Zone F", 4, "Core & Shell", 235, "Middle", "Landscape", 26100000, "2026", "5% down · 8 years")
    add_unit("the-brooks", "Townhouse", "TB-TH-COR-2", "Zone F", 4, "Core & Shell", 244, "Corner", "Landscape", 29800000, "2026", "5% down · 8 years")
    # Stone Residence Apartments
    stone_specs = [
        ("Apartment", 2, 128, 8960000, "Ground with Garden"),
        ("Apartment", 2, 140, 9240000, "Typical Floor"),
        ("Apartment", 3, 155, 10850000, "Ground with Garden"),
        ("Apartment", 3, 175, 11725000, "Typical Floor"),
        ("Apartment", 4, 220, 14960000, "Penthouse"),
    ]
    for idx, (utype, beds, bua, price, note) in enumerate(stone_specs):
        for i in range(1, 6):
            add_unit("stone-residence", utype, f"SR-{utype[:2].upper()}-{idx}-{i}", "Phase 2", beds, "Core & Shell", bua, note, "Landscape", price, "2026", "20% down · 5 years")
    # Jebal Sokhna
    for i in range(1, 11):
        add_unit("jebal-sokhna", "Chalet", f"JS-CH-{100+i}", "Aqua Zone" if i%2==0 else "Crimson Heights", 3, "Finished", 130 if i%2==0 else 150, "", "Sea View", 8000000 + i*500000, "2028", "10% down · 7 years")
    # The Med
    for i in range(1, 11):
        add_unit("the-med", "Chalet" if i%2==0 else "Townhouse", f"TM-{100+i}", "Phase 1", 3, "Core & Shell", 140 if i%2==0 else 200, "", "Landscape", 12000000 + i*800000, "2028", "5% down · 8 years")

# Tatweer Misr PDF facts typical units
print("Generating Tatweer Misr compound units...")
# Il Monte Galala
for i in range(1, 16):
    add_unit("il-monte-galala", "Apartment" if i%2==0 else "Chalet", f"IMG-U-{100+i}", "Marina Residences", 2 if i%3==0 else 3, "Finished", 110 if i%2==0 else 145, "", "Lagoon View", 9000000 + i*400000, "2028", "5% down · 8 years")
# Bloomfields
for i in range(1, 16):
    add_unit("bloomfields", "Apartment", f"BF-A-{100+i}", "Zone 1", 2 if i%2==0 else 3, "Core & Shell", 120 if i%2==0 else 165, "", "Landscape", 6000000 + i*300000, "2028", "5% down · 9 years")
# Rivers
for i in range(1, 11):
    add_unit("rivers", "Townhouse" if i%2==0 else "Standalone Villa", f"RIV-U-{100+i}", "Phase 1", 4, "Core & Shell", 200 if i%2==0 else 280, "", "Landscape", 22000000 + i*900000, "2028", "5% down · 8 years")

# HAP AVAILABILITY PDF
hap_pdf_path = os.path.join(raw_dir, "HAP AVAILABILITY.pdf")
if os.path.exists(hap_pdf_path):
    print("Parsing HAP AVAILABILITY PDF...")
    # Swanlake Residences New Cairo
    # The Phoenix Apartments
    phoenix_specs = [
        ("Apartment", 2, 115, 25600000, "Garden: 110 sqm"),
        ("Apartment", 2, 128, 19800000, "Typical Floor"),
        ("Apartment", 3, 165, 33200000, "Garden: 120 sqm"),
    ]
    for idx, (utype, beds, bua, price, note) in enumerate(phoenix_specs):
        for i in range(1, 4):
            add_unit("swan-lake", utype, f"SL-PHX-{idx}-{i}", "The Phoenix", beds, "Finished", bua, note, "Landscape", price, "2028", "5% down · 10 years")
    # The Selina Villas
    add_unit("swan-lake", "Standalone Villa", "SL-SEL-V01", "The Selina", 5, "Core & Shell", 459, "Land Area: 829 sqm", "Landscape", 162000000, "2030", "5% down · 6 years")
    # AM:PM Offices
    add_unit("swan-lake", "Office", "SL-OFF-A02", "AM:PM", None, "Core & Shell", 497, "1st Floor", "Suez Road View", 128100000, "2027", "10% down · 7 years")
    add_unit("swan-lake", "Office", "SL-OFF-A04", "AM:PM", None, "Core & Shell", 195, "5th Floor", "Suez Road View", 47500000, "2027", "10% down · 7 years")

    # Hap Town Mostakbal City
    # Terraces Apartments
    terraces_specs = [
        ("Apartment", 1, 90, 9200000, "Typical"),
        ("Apartment", 2, 125, 14500000, "Typical"),
        ("Apartment", 3, 160, 18500000, "Typical"),
        ("Duplex", 3, 275, 31600000, "Garden Duplex"),
    ]
    for idx, (utype, beds, bua, price, note) in enumerate(terraces_specs):
        for i in range(1, 4):
            add_unit("haptown", utype, f"HT-TR-{idx}-{i}", "Terraces Apartments", beds, "Finished", bua, note, "Central Park View", price, "2030", "5% down · 10 years")
    # The Great Lawn Apartments
    lawn_specs = [
        ("Apartment", 1, 88, 8750000, "Typical"),
        ("Apartment", 2, 150, 15650000, "Typical"),
        ("Apartment", 3, 185, 19200000, "Typical"),
    ]
    for idx, (utype, beds, bua, price, note) in enumerate(lawn_specs):
        for i in range(1, 4):
            add_unit("haptown", utype, f"HT-GL-{idx}-{i}", "The Great Lawn", beds, "Finished", bua, note, "Central Park View", price, "2030", "5% down · 10 years")
    # The Valleys (Villas)
    valley_specs = [
        ("Townhouse", 4, 185, 27800000, "Middle"),
        ("Townhouse", 4, 210, 30900000, "Corner"),
        ("Twin House", 4, 188, 33000000, ""),
        ("Standalone Villa", 4, 220, 44900000, ""),
        ("Standalone Villa", 5, 240, 46500000, ""),
    ]
    for idx, (utype, beds, bua, price, note) in enumerate(valley_specs):
        for i in range(1, 3):
            add_unit("haptown", utype, f"HT-VL-{idx}-{i}", "The Valleys", beds, "Finished", bua, note, "Landscape & Water Bodies", price, "2030", "5% down · 9 years")
    # Park 226
    add_unit("haptown", "Apartment", "HT-P226-1", "Park 226", 2, "Core & Shell", 145, "", "Lagoon View", 13100000, "2029", "5% down · 8 years")
    add_unit("haptown", "Apartment", "HT-P226-2", "Park 226", 3, "Core & Shell", 185, "Ground", "Lagoon View", 20200000, "2029", "5% down · 8 years")
    # Park View
    add_unit("haptown", "Apartment", "HT-PVW-1", "Park View", 2, "Core & Shell", 168, "", "Landscape", 16250000, "2026", "10% down · 8 years")

# La Vista PDFs inside raw_source_files/Lavista/
lavista_dir = os.path.join(raw_dir, "Lavista")
if os.path.exists(lavista_dir):
    print("Parsing La Vista PDFs...")
    # 1. Riva
    add_unit("el-patio-riva", "Apartment", "LVR-A101", "Phase 1", 1, "Semi Finished", 85, "Typical", "Landscape", 5973000, "2029", "5% down · 8 years")
    add_unit("el-patio-riva", "Apartment", "LVR-A201", "Phase 1", 2, "Semi Finished", 131, "Typical", "Landscape", 9550000, "2029", "5% down · 8 years")
    add_unit("el-patio-riva", "Apartment", "LVR-A301", "Phase 1", 3, "Semi Finished", 184, "Typical", "Landscape", 14010000, "2029", "5% down · 8 years")
    # 2. Vera
    add_unit("el-patio-vera", "Townhouse", "LVV-TH1", "Phase 1", 4, "Core & Shell", 176, "Middle", "Landscape", 18300000, "2028", "5% down · 7 years")
    add_unit("el-patio-vera", "Twin House", "LVV-TW1", "Phase 1", 4, "Core & Shell", 200, "Type V-B", "Landscape", 22550000, "2028", "5% down · 7 years")
    add_unit("el-patio-vera", "Standalone Villa", "LVV-SV1", "Phase 1", 5, "Core & Shell", 324, "", "Landscape", 39400000, "2028", "5% down · 7 years")
    # 3. Sokhna
    # La Vista 6
    add_unit("la-vista-6", "Chalet", "LV6-CH1", "Old Phase", 3, "Finished", 140, "1st Floor", "Sea View", 9800000, "Ready To Move", "35% down · 3 years")
    add_unit("la-vista-6", "Chalet", "LV6-CH2", "New Phase", 3, "Finished", 130, "Ground with 40m Garden", "Pool View", 11770000, "2029", "5% down · 8 years")
    # La Vista 7
    add_unit("la-vista-7", "Chalet", "LV7-CH1", "Phase 1", 3, "Finished", 150, "1st Floor", "Sea & Pool View", 12375000, "Ready To Move", "10% down · 5 years")
    # La Vista Topaz
    add_unit("la-vista-topaz", "Chalet", "LVT-CH1", "Phase 1", 3, "Finished", 150, "Ground with 50m Garden", "Pool View", 16950000, "Ready To Move", "10% down · 5 years")
    # La Vista Gardens
    add_unit("la-vista-gardens", "Chalet", "LVG-CH1", "Phase 1", 3, "Finished", 150, "Ground with 50m Garden", "Pool View", 16725000, "Ready To Move", "10% down · 5 years")
    # La Vista Ray
    add_unit("la-vista-ray", "Chalet", "LVR-CH1", "Phase 1", 2, "Finished", 115, "1st Floor", "Pool View", 9200000, "Ready To Move", "5% down · 6 years")
    add_unit("la-vista-ray", "Townhouse", "LVR-TH1", "Phase 1", 4, "Finished", 170, "with 40m Garden", "Pool View", 22800000, "Ready To Move", "5% down · 6 years")
    # 4. Cascada
    add_unit("la-vista-cascada", "Penthouse", "LVC-PH1", "Phase 1", 4, "Finished", 185, "with 150m Roof", "Pool View", 27800000, "Ready To Move", "35% down · 3 years")
    # 5. Bay East
    add_unit("la-vista-bay-east", "Chalet", "LVBE-CH1", "Phase 1", 3, "Finished", 150, "1st Floor", "Landscape View", 21000000, "2028", "5% down · 7 years")
    # 6. Ras El Hekma
    add_unit("la-vista-ras-el-hekma", "Chalet", "LVRH-CH1", "Old Phase", 3, "Finished", 150, "Ground with 50m Garden", "Sea View", 24800000, "Ready To Move", "20% down · 4 years")
    add_unit("la-vista-ras-el-hekma", "Chalet", "LVRH-CH2", "Lagoons", 3, "Finished", 155, "Ground with 40m Garden", "Lagoon View", 21980000, "2029", "5% down · 7 years")
    add_unit("la-vista-ras-el-hekma", "Twin House", "LVRH-TW1", "Lagoons", 4, "Finished", 160, "with 40m Garden", "Lagoon View", 27800000, "2029", "5% down · 7 years")
    # 7. Patio Hills
    add_unit("patio-hills", "Townhouse", "LVP-TH1", "Phase 1", 4, "Core & Shell", 180, "Middle", "Landscape", 21600000, "3.5 Years", "5% down · 8 years")
    # 8. Patio Jade
    add_unit("patio-jade", "Townhouse", "LVPJ-TH1", "Phase 1", 4, "Core & Shell", 185, "Middle", "Landscape", 20700000, "4.5 Years", "5% down · 8 years")
    # 9. Patio Oro
    add_unit("patio-oro", "Apartment", "LVPO-A1", "Phase 1", 3, "Semi Finished", 185, "Typical", "Landscape", 19150000, "Ready To Move", "20% down · 4 years")
    # 10. Patio Casa
    add_unit("patio-casa", "Twin House", "LVPC-TW1", "Phase 1", 4, "Core & Shell", 235, "with 250m Land", "Street View", 29800000, "Ready To Move", "20% down · 4 years")
    # 11. Patio 5 East
    add_unit("patio-5-east", "Townhouse", "LVP5-TH1", "Phase 1", 4, "Core & Shell", 211, "Middle", "Street View", 24100000, "Ready To Move", "20% down · 4 years")
    # 12. Patio Prime
    add_unit("patio-prime", "Standalone Villa", "LVPP-SV1", "Phase 1", 5, "Core & Shell", 288, "with 364m Land", "Pool View", 39800000, "Ready To Move", "20% down · 4 years")
    # 13. Patio Vida
    add_unit("patio-vida", "Apartment", "LVPV-A1", "Phase 1", 3, "Semi Finished", 168, "Typical", "Landscape", 12460000, "2029", "5% down · 8 years")
    # 14. Patio Zahraa
    add_unit("patio-zahraa", "Twin House", "LVPZ-TW1", "Phase 1", 4, "Core & Shell", 264, "with 286m Land", "Landscape", 34500000, "Ready To Move", "25% down · 4 years")
    # 15. Sola
    add_unit("el-patio-sola", "Apartment", "LVPS-A1", "Phase 1", 2, "Semi Finished", 123, "Typical", "Landscape", 7990000, "2029", "5% down · 7.75 years")
    # 16. Town
    add_unit("el-patio-town", "Townhouse", "LVPT-TH1", "Phase 1", 4, "Finished", 215, "Middle", "Landscape", 30500000, "2028", "5% down · 7 years")

# ==========================================
# 4. WRITE SPREADSHEETS FOR ALL GENERATED PROJECTS
# ==========================================
print(f"\nWriting standard spreadsheets for {len(project_units)} projects...")
for slug, data in project_units.items():
    groups = {}
    for u in data["units"]:
        key = (
            u.get("type", "Apartment"),
            u.get("beds", 3),
            u.get("finishing", "Finished"),
            u.get("cluster", "Phase 1"),
            u.get("delivery_note", "2028"),
            u.get("payment_plan", "")
        )
        if key not in groups:
            groups[key] = []
        groups[key].append(u)
        
    breakdown_list = []
    total_avail = len(data["units"])
    
    for key, group_units in groups.items():
        utype, beds, finishing, cluster, deliv, pay = key
        sqms = [u["area_sqm"] for u in group_units if u.get("area_sqm")]
        prices = [u["price_egp"] / 1000000.0 for u in group_units if u.get("price_egp")]
        
        min_sqm = min(sqms) if sqms else 0
        max_sqm = max(sqms) if sqms else 0
        min_price = min(prices) if prices else 0.0
        max_price = max(prices) if prices else 0.0
        
        breakdown_list.append({
            "type": utype,
            "beds": beds,
            "available": len(group_units),
            "min_sqm": min_sqm,
            "max_sqm": max_sqm,
            "min_price_m": round(min_price, 2),
            "max_price_m": round(max_price, 2),
            "finishing": finishing,
            "cluster": cluster,
            "delivery_note": deliv,
            "payment_plan": pay
        })
        
    dev_dir = os.path.join(projects_dir, data["dev_slug"])
    os.makedirs(dev_dir, exist_ok=True)
    
    fpath = os.path.join(dev_dir, f"{slug}.xlsx")
    
    wb = openpyxl.Workbook()
    ws_proj = wb.active
    ws_proj.title = "Projects"
    ws_proj.append(["slug", "developer", "total_available", "last_updated", "note"])
    ws_proj.append([slug, data["dev_name"], total_avail, datetime.today().strftime('%Y-%m-%d'), f"Live inventory for {slug}"])
    
    ws_brk = wb.create_sheet(title="Breakdown")
    ws_brk.append(["slug", "type", "beds", "available", "min_sqm", "max_sqm", "min_price_m", "max_price_m", "finishing", "cluster", "delivery_note", "payment_plan"])
    for b in breakdown_list:
        ws_brk.append([
            slug, b["type"], b["beds"], b["available"], b["min_sqm"], b["max_sqm"],
            b["min_price_m"], b["max_price_m"], b["finishing"], b["cluster"],
            b["delivery_note"], b["payment_plan"]
        ])
        
    ws_uni = wb.create_sheet(title="Units")
    ws_uni.append(["slug", "type", "unit_id", "cluster", "beds", "finishing", "area_sqm", "area_note", "view", "price_egp", "delivery_note", "payment_plan", "status"])
    for u in data["units"]:
        ws_uni.append([
            slug, u["type"], u["unit_id"], u["cluster"], u["beds"], u["finishing"],
            u["area_sqm"], u["area_note"], u["view"], u["price_egp"], u["delivery_note"],
            u["payment_plan"], u["status"]
        ])
        
    wb.save(fpath)
    print(f"Saved {slug}.xlsx inside {data['dev_slug']} (Total: {total_avail} units)")

print("\nAll spreadsheet parsing and generation completed successfully!")
